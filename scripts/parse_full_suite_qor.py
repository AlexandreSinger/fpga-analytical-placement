#!/usr/bin/env python3
import argparse
from collections import OrderedDict

from pathlib import Path

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import re
import os
import pandas as pd

METRICS = [
    "post_fl_hpwl",
    "post_dp_hpwl",
    "total_wirelength",
    "post_fl_cpd",
    "post_dp_cpd",
    "crit_path_delay",
    "route_runtime",
    "total_runtime",
]

SUMMARY_METRICS = [
    "total_wirelength",
    "crit_path_delay",
    "total_runtime",
]

def parse_args():
    parser = argparse.ArgumentParser(
            description="TODO"
    )

    parser.add_argument(
        "task_list_file_path",
        type=str,
    )

    parser.add_argument(
        "-o",
        "--output_spreadsheet",
        metavar="OUTPUT.XLSX",
        default="output.xlsx",
    )

    return parser.parse_args()

def parse_task_list(task_list: str) -> list[str]:
    if not os.path.isfile(task_list):
        raise ValueError("Task list is not a file")

    # NOTE: We assume that the task list is in the same directory as the tasks.
    task_dirs = []
    with open(task_list, 'r') as f:
        for line in f.readlines():
            task_dirs.append(os.path.join(os.path.dirname(task_list), line.strip()))

    return task_dirs

def get_latest_run_folder(base_dir: str) -> Path:
    assert(os.path.isdir(base_dir))
    base_dir = Path(base_dir)
    run_re = re.compile(r"^run(\d+)$")
    runs = []
    for entry in base_dir.iterdir():
        if entry.is_dir():
            m = run_re.match(entry.name)
            if m:
                runs.append((int(m.group(1)), entry))

    assert(len(runs) != 0)

    return max(runs, key=lambda x: x[0])[1]


def get_qor_results_for_tasks(task_dirs: list[str]) -> list[tuple[str, Path]]:
    qor_result_files = []
    for task_dir in task_dirs:
        assert(os.path.isdir(task_dir))
        run_folder = get_latest_run_folder(task_dir)
        qor_file = os.path.join(run_folder, "qor_results.txt")
        assert(os.path.isfile(qor_file))
        qor_result_files.append((os.path.basename(task_dir), qor_file))

    return qor_result_files

def make_ratios_and_summary(ratio_sheet, summary_sheet, raw_sheets):
    # Get the task base-names (removing the no_ap / ap)
    task_base_names = []
    for task_name in raw_sheets.keys():
        if not task_name.endswith("_no_ap"):
            continue
        task_base_names.append(task_name[:-6])

    # NOTE: THIS IS FRAGILE. The order is not gaurenteed!
    ratio_sheet.append(["test_suite", "arch", "circuit"] + METRICS)
    summary_sheet.append(["test_suite"] + SUMMARY_METRICS + ["Baseline Unroutes"])
    for task_base_name in task_base_names:
        task_start_row = ratio_sheet.max_row + 1
        ap_raw_sheet = raw_sheets[task_base_name + "_ap"]
        no_ap_raw_sheet = raw_sheets[task_base_name + "_no_ap"]
        for row in range(2, ap_raw_sheet.max_row + 1):
            row_data = [task_base_name]
            # Grab the arch and circuit.
            # TODO: Ensure that these are these two columns.
            row_data.append(ap_raw_sheet.cell(row=row, column=1).value.strip())
            row_data.append(ap_raw_sheet.cell(row=row, column=2).value.strip())
            # Find the wirelength cell. This is used to detect unroutes.
            ap_wirelength_cell_coord = None
            no_ap_wirelength_cell_coord = None
            for col in range(3, ap_raw_sheet.max_column + 1):
                ap_header_name = ap_raw_sheet.cell(row=1, column=col).value.strip()
                if ap_header_name != "total_wirelength":
                    continue
                ap_wirelength_cell = ap_raw_sheet.cell(row=row, column=col)
                no_ap_wirelength_cell = no_ap_raw_sheet.cell(row=row, column=col)
                ap_wirelength_cell_coord = f"'{ap_raw_sheet.title}'!{ap_wirelength_cell.coordinate}"
                no_ap_wirelength_cell_coord = f"'{no_ap_raw_sheet.title}'!{no_ap_wirelength_cell.coordinate}"
                break
            assert(ap_wirelength_cell_coord is not None)
            assert(no_ap_wirelength_cell_coord is not None)
            # Get the other metrics.
            for col in range(3, ap_raw_sheet.max_column + 1):
                ap_header_name = ap_raw_sheet.cell(row=1, column=col).value.strip()
                no_ap_header_name = no_ap_raw_sheet.cell(row=1, column=col).value.strip()
                assert ap_header_name == no_ap_header_name, "Header values must match"

                if ap_header_name not in METRICS:
                    continue

                ap_cell = ap_raw_sheet.cell(row=row, column=col)
                no_ap_cell = no_ap_raw_sheet.cell(row=row, column=col)
                ap_cell_coord = f"'{ap_raw_sheet.title}'!{ap_cell.coordinate}"
                no_ap_cell_coord = f"'{no_ap_raw_sheet.title}'!{no_ap_cell.coordinate}"
                row_data.append(f"=IF(OR({ap_wirelength_cell_coord}=-1, {no_ap_wirelength_cell_coord}=-1), \"\", {ap_cell_coord}/{no_ap_cell_coord})")
                # if ap_value == -1 or no_ap_value == -1:
                #     row_data.append("")
                # else:
                #     row_data.append(ap_value / no_ap_value)
            ratio_sheet.append(row_data)
        task_end_row = ratio_sheet.max_row

        # Compute the geomean
        geomean_row_data = [task_base_name, "", "GEOMEAN"]
        for col in range(4, ratio_sheet.max_column + 1):
            start_coordinate = ratio_sheet.cell(row=task_start_row, column=col).coordinate
            end_coordinate = ratio_sheet.cell(row=task_end_row, column=col).coordinate
            geomean_row_data.append(f"=GEOMEAN({start_coordinate}:{end_coordinate})")
        ratio_sheet.append(geomean_row_data)

        # Get the number of unroutes
        unroute_count_row_data = [task_base_name, "", "UNROUTES"]
        for col in range(4, ratio_sheet.max_column + 1):
            if ratio_sheet.cell(row=1, column=col).value.strip() != "total_wirelength":
                continue
            start_coordinate = ratio_sheet.cell(row=task_start_row, column=col).coordinate
            end_coordinate = ratio_sheet.cell(row=task_end_row, column=col).coordinate
            unroute_count_row_data.append(f"={task_end_row - task_start_row + 1} - COUNT({start_coordinate}:{end_coordinate})")
            break
        ratio_sheet.append(unroute_count_row_data)

        summary_row_data = [task_base_name]
        for col in range(4, ratio_sheet.max_column + 1):
            if ratio_sheet.cell(row=1, column=col).value.strip() not in SUMMARY_METRICS:
                continue
            geomean_row = ratio_sheet.max_row - 1
            geomean_cell = ratio_sheet.cell(row=geomean_row, column=col)
            geomean_cell_coord = f"'{ratio_sheet.title}'!{geomean_cell.coordinate}"
            summary_row_data.append(f"={geomean_cell_coord}")
        unroute_row = ratio_sheet.max_row
        unroute_cell = ratio_sheet.cell(row=unroute_row, column=4)
        summary_row_data.append(f"='{ratio_sheet.title}'!{unroute_cell.coordinate}")
        summary_sheet.append(summary_row_data)

        # Add a space after each task for readability
        ratio_sheet.append([""])



def main():
    args = parse_args()

    task_dirs = parse_task_list(args.task_list_file_path)

    qor_result_files = get_qor_results_for_tasks(task_dirs)

    wb = openpyxl.Workbook()

    # Remove the default sheet
    wb.remove(wb.active)

    # Load all the raw data into separate sheets
    raw_sheets = OrderedDict()
    for (task_name, qor_result_file) in qor_result_files:
        print(f"Loading: {qor_result_file}")
        df = pd.read_csv(qor_result_file, skipinitialspace=True, sep="\t", engine="python")

        # Convert to work sheet
        ws = wb.create_sheet(title=task_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        raw_sheets[task_name] = ws

    # Create sheet of ratios
    ratio_sheet = wb.create_sheet("ratios", index=0)
    summary_sheet = wb.create_sheet("summary", index=0)
    make_ratios_and_summary(ratio_sheet, summary_sheet, raw_sheets)


    wb.save(args.output_spreadsheet)


if __name__ == "__main__":
    main()
